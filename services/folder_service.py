
import os
import pytesseract
import fitz  # PyMuPDF para trabajar con PDFs
from PIL import Image, ImageEnhance, ImageFilter
import pandas as pd
import numpy as np
import services.func_extrac_data as look_data  # Asegúrate de que esta ruta sea correcta
import utils.app_logger as log
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sys

from PyQt5.QtCore import QObject, pyqtSignal, QThread

import cv2

# Diccionario para traducir meses entre inglés y español
months_translator = {
    "January": "Enero", "February": "Febrero", "March": "Marzo", "April": "Abril",
    "May": "Mayo", "June": "Junio", "July": "Julio", "August": "Agosto",
    "September": "Septiembre", "October": "Octubre", "November": "Noviembre", "December": "Diciembre",
    "Enero": "January", "Febrero": "February", "Marzo": "March", "Abril": "April",
    "Mayo": "May", "Junio": "June", "Julio": "July", "Agosto": "August",
    "Septiembre": "September", "Octubre": "October", "Noviembre": "November", "Diciembre": "December"
}

months_numbers = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12,
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
    "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8,
    "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
}


class FolderProcessor(QObject):
    progressChanged = pyqtSignal(int)
    
    def __init__(self, output_folder, output_file):
        super().__init__()  # Llamar al constructor de QObject
        self.output_folder = os.path.join(output_folder)  # Carpeta donde se guardarán los archivos
        self.output_file = output_file  # Archivo Excel final

        # self.carpeta_ejecutable = os.path.dirname(sys.executable)
        self.carpeta_ejecutable = os.path.dirname(os.path.abspath(__file__))
        self.folder_output_ejecutable_unificado = os.path.join(self.carpeta_ejecutable, "output")
        # self.folder_output_ejecutable_clientes = os.path.join(self.carpeta_ejecutable, "carpeta_archivos_clientes")

        os.makedirs(self.folder_output_ejecutable_unificado, exist_ok=True)  # Crear carpeta de salida si no existe
        # os.makedirs(self.folder_output_ejecutable_clientes, exist_ok=True)  # Crear carpeta de salida si no existe
        os.makedirs(output_folder, exist_ok=True)  # Crear carpeta de salida si no existe

    def process(self, parent_folder, year, month):
        year_folder_name = str(year)
        translated_month = self.translate_month(month)  # Traduce el mes al idioma necesario
        month_number = self.get_month_number(month)
        combined_df = pd.DataFrame()

        client_folders = [f for f in os.listdir(parent_folder) if os.path.isdir(os.path.join(parent_folder, f))]
        total_folders = len(client_folders)
        total_files = 0
        self.processed_files = 0

        # Contar todos los archivos que se van a procesar
        for client_folder in client_folders:
            client_path = os.path.join(parent_folder, client_folder)
            target_path = self._get_target_path(client_path, year_folder_name, translated_month, month_number)
            if not target_path:
                log.log_info(f"No se encontró carpeta válida en: {os.path.abspath(client_path)}")
                continue
            total_files += len([f for f in os.listdir(target_path) if any(f.endswith(suffix + ".pdf") for suffix in ["EDD", "941", f"{year}"])])

        for client_folder in client_folders:
            client_path = os.path.join(parent_folder, client_folder)
            target_path = self._get_target_path(client_path, year_folder_name, translated_month, month_number)
            if not target_path:
                # print(f"No se encontró carpeta válida en: {os.path.abspath(client_path)}")
                continue

            print(f"Procesando carpeta: {os.path.abspath(target_path)}")
            df_weekly = self.process_weekly_files(target_path, year, translated_month, total_files, self.processed_files)

            #! Prueba de excel por cliente
            # carpeta_cliente_nom = os.path.basename(os.path.dirname(os.path.dirname(target_path)))
            # df_cliente = self.prepare_data(df_weekly)
            # self.save_to_excel(os.path.join(self.folder_output_ejecutable_clientes, f"{carpeta_cliente_nom}.xlsx" ), df_cliente)
            #! Prueba de excel por cliente

            combined_df = pd.concat([combined_df, df_weekly], ignore_index=True)
            

        df_modify = self.prepare_data(combined_df)
        # Guardar todos los datos combinados en un solo archivo Excel
        self.save_to_excel(os.path.join(self.folder_output_ejecutable_unificado, self.output_file), df_modify)

    def clean_path_segment(self, segment):
        # Reemplazar espacios en "mes-numero" para que quede "numero_mes+mes"
        return segment.replace(" ", "")
    
    def _get_target_path(self, path, year, month, number):

        year_folder = None
        
        for root, dirs, files in os.walk(path):
            for name in dirs:
                if str(year) in name and "Payroll" in name:
                    year_folder = os.path.join(root, name)
                elif str(year) in name and "payroll" in name:
                    year_folder = os.path.join(root, name)
                elif str(year) in name:
                        year_folder = os.path.join(root, name)

        if year_folder is not None: 
            for root, dirs, files in os.walk(year_folder):
                for name in dirs:
                    if str(number) in name and month in name:
                        return os.path.join(root, name)
                    elif str(number) in name:
                        return os.path.join(root, name)
        return None


    def process_weekly_files(self, folder_path, year, month, total_files, processed_files):
        columnas = ['tipo_archivo', 'fecha_pdf', 'Name', 'federal_tax_941', 'state_tax_edd',
                        '941_payment_amount', 'EDD_payment_amount', 'account_number', 'date_pay_settle', 'carpeta_cliente']
        df = pd.DataFrame(columns=columnas)
        carpeta_cliente = os.path.basename(os.path.dirname(os.path.dirname(folder_path)))
        print(carpeta_cliente)

        week_read = None
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)

            if not os.path.isfile(file_path):
                continue

            # Verificar si los primeros 8 dígitos del nombre del archivo son numéricos
            if not file_name[:8].isdigit():
                continue

            # Verificar la longitud del nombre del archivo sin la extensión
            nombre_sin_extension = os.path.splitext(file_name)[0]
            if len(nombre_sin_extension) > 12:
                continue

            if not any(file_name.endswith(suffix + ".pdf") for suffix in ["EDD", "941", f"{year}"]):
                continue

            if any(file_name.endswith(suffix + ".pdf") for suffix in ["EDD", "941"]):
                caracter_nueve = file_name[8]
                if caracter_nueve != ' ':
                    continue
            elif any(file_name.endswith(suffix + ".pdf") for suffix in [f"{year}"]):
                caracter_nueve = file_name[8]
                if caracter_nueve != '.':
                    continue
            
            # Validamos si hay archivos faltantes para la semana que estamos recorriendo
            
            week = file_name[:8]
            look_files = (f"{week} EDD.pdf", f"{week} 941.pdf", f"{week}.pdf")
            if week_read != week:
                for file in look_files: # Ruta completa del archivo file_path = os.path.join(folder_path, file_name)
                    file_possible_path = os.path.join(folder_path, file)
                    if not os.path.isfile(file_possible_path): 
                        # print(f"El archivo {file} no existe.")
                        log.log_info(f"El archivo {file} no existe en la carpeta {folder_path}")
                        tipo_archivo = file.replace(f"{week} ", "").replace(".pdf", "")
                        tipo_archivo = 'General' if tipo_archivo == week else tipo_archivo
                        df_no_files = pd.DataFrame([{
                            'tipo_archivo': tipo_archivo,
                            'fecha_pdf': week,
                            'carpeta_cliente': carpeta_cliente,
                            'ruta_archivo' : "Archivo no encontrado"
                        }])
                        # Agregar las columnas específicas dependiendo del tipo de archivo
                        if tipo_archivo == "EDD":
                            df_no_files['EDD_payment_amount'] = "Archivo no encontrado"
                        elif tipo_archivo == "941":
                            df_no_files['941_payment_amount'] = "Archivo no encontrado"
                        elif tipo_archivo == "General":
                            df_no_files['federal_tax_941'] = "Archivo no encontrado"
                            df_no_files['state_tax_edd'] = "Archivo no encontrado"


                        df = pd.concat([df, df_no_files], ignore_index=True)
                        week_read = week

            print(f"Procesando archivo: {os.path.abspath(file_path)}")

            text = self.process_file_with_ocr(file_path)
            with open(os.path.join(folder_path, f'{file_name.replace('.pdf','.txt')}'), 'w', encoding='utf-8',) as txt_file:
                txt_file.write(text)
            
            datos = self.handle_extracted_data(file_name, text, carpeta_cliente, file_path,month, year)

            df = pd.concat([df, datos], ignore_index=True)

            # Actualizar el progreso y emitir la señal
            self.processed_files += 1
            progress = int((self.processed_files / total_files) * 100)
            self.progressChanged.emit(progress)

        return df


    def process_file_with_ocr(self, file_path):
        try:
            pdf_document = fitz.open(file_path)
            extracted_text = ""
            last_page_number = len(pdf_document) - 1
            page = pdf_document[last_page_number]
            pix = page.get_pixmap(dpi=300)
            image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            extracted_text += pytesseract.image_to_string(image, lang='spa', config='--dpi 300 --psm 6')
            return extracted_text
        except Exception as e:
            print(f"Error al procesar el archivo con OCR: {e}")
            return ""

    def handle_extracted_data(self, file_name, text, carpeta_cliente, file_path, month, year):
        def format_date(date_str):
            if date_str is None:
                return "0"
            
            try:
                return pd.to_datetime(date_str).strftime('%Y-%m-%d')
            except ValueError:
                return date_str

        if file_name.endswith("941.pdf"):
            return pd.DataFrame([{
                'tipo_archivo': "941",
                'fecha_pdf': file_name.replace('941.pdf', ''),
                'Name': look_data.extract_payer_name(text),
                '941_payment_amount': look_data.extract_payment_amount_941(text),
                'account_number': look_data.extract_account_number(text),
                'date_pay_settle': format_date(look_data.extract_settlement_date(text)),
                'carpeta_cliente': carpeta_cliente,
                'ruta_archivo' : file_path
            }])
        elif file_name.endswith("EDD.pdf"):
            return pd.DataFrame([{
                'tipo_archivo': "EDD",
                'fecha_pdf': file_name.replace('EDD.pdf', ''),
                'Name': look_data.extract_name(text),
                'EDD_payment_amount': look_data.extract_payment_amount_edd(text, 2),
                'account_number': look_data.extract_account_number(text),
                'date_pay_settle': format_date(look_data.extract_payment_date(text)),
                'carpeta_cliente': carpeta_cliente,
                'ruta_archivo' : file_path
            }])
        else:
            return pd.DataFrame([{
                'tipo_archivo': "general",
                'fecha_pdf': file_name.replace('.pdf', ''),
                'Name': look_data.extract_company_name(text),
                'federal_tax_941': look_data.extract_payment_amount_general_941(text),
                'state_tax_edd': look_data.extract_payment_amount_general_edd(text),
                'carpeta_cliente': carpeta_cliente,
                'ruta_archivo' : file_path
            }])

    def group_by_weekly(self, dataframe):
        """
        Agrupa los datos por carpeta_cliente y fecha_pdf para consolidar información.
        """

        try:
            dataframe['Check date'] = dataframe['Check date'].astype(int)
            # Obtener combinaciones únicas de 'carpeta_cliente' y 'fecha_pdf'
            combinaciones_unicas = dataframe[['Company', 'Check date']].drop_duplicates()
        except Exception as e:
            print("Ocurrio un error al intentar parsear la columnas Check Date a int()")
        
        resultados = []
        
        for _, fila in combinaciones_unicas.iterrows():
            carpeta_cliente = fila['Company']
            fecha = fila['Check date']
            df_filtrado = dataframe[(dataframe['Company'] == carpeta_cliente) & (dataframe['Check date'] == fecha)]
            resultado = {'Company': carpeta_cliente, 'Check date': fecha}
            
            for columna in ['Federal Tax', 'State Tax', 'Payment date' ,'941', 'EDD']:
                valores = df_filtrado[columna].replace('', np.nan).dropna()
                resultado[columna] = valores.iloc[0] if not valores.empty else np.nan
            
            resultados.append(resultado)

        return pd.DataFrame(resultados)

    def format_excel_headers(self, file_path, headers):
        """
        Formatea los encabezados de un archivo Excel con las especificaciones proporcionadas.
        
        :param file_path: Ruta al archivo Excel.
        :param headers: Lista de encabezados que se esperan en la primera fila.
        """
        try:
            # Cargar el archivo Excel
            wb = load_workbook(file_path)
            ws = wb["Resumen"]

            # Configuración de estilo
            header_style = {
                "font": Font(name="Calibri", size=20, bold=False, color="000000"),
                "fill": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            }

            # Aplicar formato a los encabezados
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                for attr, value in header_style.items():
                    setattr(cell, attr, value)

            # Ajustar automáticamente el ancho de las columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Obtiene la letra de la columna
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Guardar los cambios en el archivo
            wb.save(file_path)
            print(f"Encabezados formateados correctamente en el archivo: {file_path}")

        except Exception as e:
            print(f"Error al formatear los encabezados: {e}")


    def crossamounts(self, path_file):
        """
        Inserta fórmulas en las columnas Balance 941 y Balance EDD.
        """
        # Seleccionar archivo Excel
        excel_path = path_file

        # Abrir el archivo Excel con openpyxl
        try:
            wb = load_workbook(excel_path)
            ws = wb['Resumen']

            ws.cell(1,8).value = 'Balance 941'
            ws.cell(1,9).value = 'Balance EDD'
            
            # Recorrer las filas del DataFrame
            for row in range(2, ws.max_row + 1):  # Asume que la fila 1 es el encabezado
                # Verificar las celdas en las columnas C, D, F y G
                if (ws[f"C{row}"].value not in ["Archivo no encontrado", "No se pudo obtener debido al formato del archivo"] and
                    ws[f"F{row}"].value not in ["Archivo no encontrado", "No se pudo obtener debido al formato del archivo"]):
                    # Si las celdas no contienen esos textos, agregar las fórmulas
                    ws[f"H{row}"] = f"=C{row}-F{row}"  # Balance 941
            
                        # Recorrer las filas del DataFrame
            for row in range(2, ws.max_row + 1):  # Asume que la fila 1 es el encabezado
                # Verificar las celdas en las columnas C, D, F y G
                if (ws[f"D{row}"].value not in ["Archivo no encontrado", "No se pudo obtener debido al formato del archivo"] and
                    ws[f"G{row}"].value not in ["Archivo no encontrado", "No se pudo obtener debido al formato del archivo"]):
                    # Si las celdas no contienen esos textos, agregar las fórmulas
                    ws[f"I{row}"] = f"=D{row}-G{row}"  # Balance EDD

            # Guardar cambios
            wb.save(excel_path)
            print("Fórmulas ingresadas en el archivo Excel.")
        except Exception as e:
            print("Error al cargar la formula")
    

    def validation_data(self, path_file):
        """
        Indicamos resaltando con un color que archivos no fueron encontrados en sus carpetas
        """
        # Seleccionar archivo Excel
        excel_path = path_file

        # Abrir el archivo Excel con openpyxl
        try:
            wb = load_workbook(excel_path)
            ws = wb['Datos combinados']

            # Definir el color de fondo (amarillo)
            color_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            color_fill_format_data = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            # Insertar fórmulas en las columnas H y I
            for row in range(2, ws.max_row + 1):  # Asume que la fila 1 es el encabezado
                if (ws[f"D{row}"].value == "No se pudo obtener debido al formato del archivo" or
                    ws[f"E{row}"].value == "No se pudo obtener debido al formato del archivo" or
                    ws[f"F{row}"].value == "No se pudo obtener debido al formato del archivo" or
                    ws[f"G{row}"].value == "No se pudo obtener debido al formato del archivo"):
                        for col in range(1, 12): # Recorrer todas las celdas en la fila
                            cell = ws.cell(row=row, column=col)
                            cell.fill = color_fill_format_data
                elif ws[f"K{row}"].value == "Archivo no encontrado":
                    for col in range(1, 12): # Recorrer todas las celdas en la fila
                        cell = ws.cell(row=row, column=col)
                        cell.fill = color_fill

            ws_resumen = wb['Resumen']
            # Recorrer las filas del DataFrame
            for row in range(2, ws_resumen.max_row + 1):  # Asume que la fila 1 es el encabezado
                # Verificar las celdas en las columnas C, D, F y G
                if(ws_resumen[f"C{row}"].value == "No se pudo obtener debido al formato del archivo" or
                    ws_resumen[f"D{row}"].value == "No se pudo obtener debido al formato del archivo" or
                    ws_resumen[f"F{row}"].value == "No se pudo obtener debido al formato del archivo" or
                    ws_resumen[f"G{row}"].value == "No se pudo obtener debido al formato del archivo"):
                    # Si las celdas no contienen esos textos, agregar las fórmulas
                    for col in range(1, 10): # Recorrer todas las celdas en la fila
                        cell = ws_resumen.cell(row=row, column=col)
                        cell.fill = color_fill_format_data
                    continue
                
                elif (ws_resumen[f"C{row}"].value == "Archivo no encontrado" or
                    ws_resumen[f"D{row}"].value == "Archivo no encontrado" or
                    ws_resumen[f"F{row}"].value == "Archivo no encontrado" or
                    ws_resumen[f"G{row}"].value == "Archivo no encontrado"):
                    # Si las celdas no contienen esos textos, agregar las fórmulas
                    for col in range(1, 10): # Recorrer todas las celdas en la fila
                        cell = ws_resumen.cell(row=row, column=col)
                        cell.fill = color_fill


            # Guardar cambios
            wb.save(excel_path)
            print("Excel modificado con exito.")
        except Exception as e:
            print("Error al identificar errores de OCR o archivos faltantes")


    def prepare_data(self, df):

        df = pd.DataFrame(df)

        df['federal_tax_941'] = df['federal_tax_941'].str.replace(',', '')
        # df['federal_tax_941'] = df['federal_tax_941'].str.replace('.', ',')

        df['state_tax_edd'] = df['state_tax_edd'].str.replace(',', '')
        # df['state_tax_edd'] = df['state_tax_edd'].str.replace('.', ',')

        df['941_payment_amount'] = df['941_payment_amount'].str.replace(',', '')
        # df['941_payment_amount'] = df['941_payment_amount'].str.replace('.', ',')

        df['EDD_payment_amount'] = df['EDD_payment_amount'].str.replace(',', '')
        # df['EDD_payment_amount'] = df['EDD_payment_amount'].str.replace('.', ',')


        df = df.rename(columns={
            'carpeta_cliente': 'Company',
            'fecha_pdf': 'Check date',
            'federal_tax_941': 'Federal Tax',
            'state_tax_edd': 'State Tax', 
            'date_pay_settle': 'Payment date', 
            '941_payment_amount': '941', 
            'EDD_payment_amount':'EDD'
            })

        return df


    def save_to_excel(self, file_path, df_new):

        try:
            # Escribir el DataFrame en una nueva hoja            
            df_resultado = self.group_by_weekly(df_new)

            # df_resultado.to_excel(ruta1, index=False, sheet_name="Resumen")
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_resultado.to_excel(writer, index=False, sheet_name='Resumen')
                df_new.to_excel(writer, index=False, sheet_name='Datos combinados')

            print(f"Datos guardados en la hoja Datos combinados del archivo {file_path}")

        except Exception as e:
            print(f"Error al guardar los datos en el archivo Excel: {e}")
        
        self.crossamounts(file_path)
        headers = ["Company","Check date","Federal Tax","State Tax","Payment date","941",'EDD', 'Balance 941', 'Balance EDD']
        self.format_excel_headers(file_path, headers)
        self.validation_data(file_path)

    def translate_month(self, month):
        return months_translator.get(month, month)
    
    def get_month_number(self, month): 
        return months_numbers.get(month, "Mes desconocido")
    
