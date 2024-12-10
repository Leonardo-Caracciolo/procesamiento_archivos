import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.sheet_name = "Resumen"
        self.required_columns = {"tipo_archivo", "fecha_pdf", "account_number"}

    def load_excel(self):
        """Carga el archivo Excel y devuelve un DataFrame de la hoja especificada."""
        try:
            excel_data = pd.ExcelFile(self.file_path)
            if self.sheet_name in excel_data.sheet_names:
                return excel_data.parse(self.sheet_name)
            else:
                print(f"La hoja '{self.sheet_name}' no existe en el archivo.")
                return None
        except Exception as e:
            print(f"Error al cargar el archivo Excel: {e}")
            return None

    def validate_columns(self, df):
        """Valida que el DataFrame tenga las columnas requeridas."""
        missing_columns = self.required_columns - set(df.columns)
        if missing_columns:
            print(f"Faltan las columnas requeridas: {missing_columns}")
            return False
        return True

    def find_discrepancies(self, df):
        """
        Encuentra las filas con discrepancias en la columna 'account_number'.
        Identifica filas cuyo 'account_number' difiera del valor más común del grupo.
        """
        discrepancies = []
        grouped = df.groupby(["fecha_pdf"])
        for _, group_df in grouped:
            if len(group_df["account_number"].unique()) > 1:
                # Identificar el valor más común en el grupo
                most_common = group_df["account_number"].mode()[0]
                # Encontrar filas con valores diferentes al más común
                discrepant_rows = group_df[group_df["account_number"] != most_common].index
                discrepancies.extend(discrepant_rows.tolist())
        return discrepancies

    def highlight_rows(self, rows_to_highlight):
        """Pinta de rojo las filas con discrepancias en la columna A usando openpyxl."""
        try:
            wb = load_workbook(self.file_path)
            ws = wb[self.sheet_name]
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for idx in rows_to_highlight:
                excel_row = idx + 2  # Ajustar índice de pandas a filas de Excel
                ws[f"A{excel_row}"].fill = red_fill
            # Guardar el archivo resaltado
            output_path = self.file_path.replace(".xlsx", "_resaltado.xlsx")
            wb.save(output_path)
            print(f"El archivo procesado se ha guardado en: {output_path}")
        except Exception as e:
            print(f"Error al resaltar las filas en el archivo Excel: {e}")

    def process(self):
        """Orquesta todas las operaciones para procesar el archivo Excel."""
        df = self.load_excel()
        if df is not None and self.validate_columns(df):
            rows_to_highlight = self.find_discrepancies(df)
            if rows_to_highlight:
                self.highlight_rows(rows_to_highlight)
            else:
                print("No se encontraron discrepancias en 'account_number'.")
        else:
            print("El archivo no pudo ser procesado.")
