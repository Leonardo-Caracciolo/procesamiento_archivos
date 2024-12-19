from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def format_excel_headers(file_path, headers):
    """
    Formatea los encabezados de un archivo Excel con las especificaciones proporcionadas.
    
    :param file_path: Ruta al archivo Excel.
    :param headers: Lista de encabezados que se esperan en la primera fila.
    """
    try:
        # Cargar el archivo Excel
        wb = load_workbook(file_path)
        ws = wb["Resumen"]

        # Configuración de estilo
        header_style = {
            "font": Font(name="Calibri", size=20, bold=True, color="000000"),
            "fill": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

        # Aplicar formato a los encabezados
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            for attr, value in header_style.items():
                setattr(cell, attr, value)

        # Ajustar automáticamente el ancho de las columnas
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(len(str(cell.value or "")) for cell in col) + 2

        # Guardar los cambios en el archivo
        wb.save(file_path)
        print(f"Encabezados formateados correctamente en el archivo: {file_path}")

    except Exception as e:
        print(f"Error al formatear los encabezados: {e}")

file_path = r"L:\Procesamiento_PDF\procesamiento_archivos\Data\Inputs\datos_combinados - copia.xlsx"
headers = ["carpeta_cliente","fecha_pdf","federal_tax_941","state_tax_edd","941_payment_amount","EDD_payment_amount"]


headers = ["Company","Check date","Federal Tax","State Tax","Payment date","941",'EDD', 'Balance 941', 'Balance EDD']

format_excel_headers(file_path, headers)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Crear un DataFrame de ejemplo
df = pd.DataFrame({
    'columna1': [1, 2, 3],
    'columna2': [4, 5, 6],
    'columna_muy_larga': [7, 8, 9]
})

# Ruta del archivo Excel
ruta_excel = 'ruta/al/archivo.xlsx'

# Guardar el DataFrame en un archivo Excel
df.to_excel(ruta_excel, index=False, sheet_name='Hoja1')

# Cargar el archivo Excel
workbook = load_workbook(ruta_excel)

# Seleccionar la hoja donde ajustar el ancho de las columnas
sheet = workbook['Hoja1']

# Ajustar el ancho de las columnas para que se muestren bien los encabezados
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter  # Obtener la letra de la columna
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column].width = adjusted_width

# Guardar los cambios en el archivo Excel
workbook.save(ruta_excel)

print("Ancho de las columnas ajustado correctamente.")


