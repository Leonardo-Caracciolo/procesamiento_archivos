
import os
import pytesseract
import fitz  # PyMuPDF para trabajar con PDFs
from PIL import Image, ImageEnhance, ImageFilter
import pandas as pd
import numpy as np
# import services.func_extrac_data as look_data  # Asegúrate de que esta ruta sea correcta
# import utils.app_logger as log
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sys
import re
from PyQt5.QtCore import QObject, pyqtSignal, QThread

def update_master_with_comparisons(file_path):
    try:
        # Cargar el archivo Excel
        wb = load_workbook(file_path)
        ws_resumen = wb['Resumen']
        ws_master = wb['Master']

        # Crear un diccionario para almacenar las sumas por cliente
        resumen_sums = {}
        for row in range(2, ws_resumen.max_row + 1):  # Asume que la fila 1 es el encabezado
            company = ws_resumen[f"A{row}"].value  # Columna 'Company'
            federal_tax = ws_resumen[f"C{row}"].value  # Columna 'Federal Tax'
            state_tax = ws_resumen[f"D{row}"].value  # Columna 'State Tax'

            # Verificamos que el campo se haya extraido del archivo PDF
            if (federal_tax in ("Archivo no encontrado", "No se pudo obtener debido al formato del archivo") 
                or state_tax in ("Archivo no encontrado", "No se pudo obtener debido al formato del archivo")):
                continue

            # Convertir valores a float si no son nulos
            federal_tax = float(federal_tax.replace(",", "").strip()) if federal_tax else 0
            state_tax = float(state_tax.replace(",", "").strip()) if state_tax else 0

            if company not in resumen_sums:
                resumen_sums[company] = {"Federal Tax": 0, "State Tax": 0}

            resumen_sums[company]["Federal Tax"] += federal_tax
            resumen_sums[company]["State Tax"] += state_tax

        # Agregar columnas adicionales en la hoja 'Master'
        if "Total Federal Resumen" not in [cell.value for cell in ws_master[1]]:
            ws_master.cell(row=1, column=ws_master.max_column + 1, value="Total Federal Resumen")
            ws_master.cell(row=1, column=ws_master.max_column + 1, value="Total State Resumen")

        col_federal_resumen = ws_master.max_column - 1
        col_state_resumen = ws_master.max_column

        # Actualizar la hoja 'Master'
        for row in range(2, ws_master.max_row + 1):  # Asume que la fila 1 es el encabezado
            company = ws_master[f"A{row}"].value  # Columna 'Company'
            total_federal_tax = ws_master[f"B{row}"].value  # Columna 'TOTAL FEDERAL TAX LIABILITY'
            total_state_tax = ws_master[f"C{row}"].value  # Columna 'TOTAL STATE TAX'

            if total_federal_tax == "Archivo no encontrado" or total_state_tax == "Archivo no encontrado":
                continue

            federal_sum = resumen_sums.get(company, {}).get("Federal Tax", 0)
            state_sum = resumen_sums.get(company, {}).get("State Tax", 0)

            # Actualizar los valores de resumen en las nuevas columnas
            ws_master.cell(row=row, column=col_federal_resumen).value = federal_sum
            ws_master.cell(row=row, column=col_state_resumen).value = state_sum

        # Crear un nuevo DataFrame con las columnas existentes y las nuevas diferencias
        data = [[cell.value for cell in row] for row in ws_master.iter_rows()]
        df = pd.DataFrame(data[1:], columns=data[0])  # Usar la primera fila como encabezados

        # Agregar las columnas de diferencia al DataFrame si no existen
        if "Diferencia Federal" not in df.columns:
            df["Diferencia Federal"] = None

        if "Diferencia State" not in df.columns:
            df["Diferencia State"] = None

        # Reordenar las columnas según los encabezados deseados
        headers = [
            "Company",
            "TOTAL FEDERAL TAX LIABILITY",
            "Total Federal Resumen",
            "Diferencia Federal",
            "TOTAL STATE TAX",
            "Total State Resumen",
            "Diferencia State"
        ]

        df = df[headers]  # Reordenar las columnas según la lista

        # Sobrescribir la hoja "Master" con las columnas reordenadas
        for col_num, header in enumerate(headers, start=1):
            ws_master.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(df.values, start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                ws_master.cell(row=row_num, column=col_num, value=cell_value)

        for row in range(2, ws_master.max_row + 1):  # Asume que la fila 1 es el encabezado
            
            fed_tax_master = ws_master[f"B{row}"].value
            state_tax_master = ws_master[f"E{row}"].value

            if (fed_tax_master in ("Archivo no encontrado", "No se pudo obtener debido al formato del archivo") 
                or state_tax_master in ("Archivo no encontrado", "No se pudo obtener debido al formato del archivo")):
                continue

            fed_tax_resumen = ws_master[f"C{row}"].value
            state_tax_resumen = ws_master[f"F{row}"].value

            ws_master[f"C{row}"].value = str(fed_tax_resumen).replace(",",".")
            ws_master[f"F{row}"].value = str(state_tax_resumen).replace(",",".")

            ws_master[f"D{row}"].value = f"=B{row}-C{row}"
            ws_master[f"G{row}"].value = f"=E{row}-F{row}"

        # Guardar los cambios
        wb.save(file_path)
        print(f"Hoja 'Master' actualizada y columnas reordenadas en {file_path}")
    except Exception as e:
        print(f"Error actualizando la hoja 'Master': {e}")

if __name__ == "__main__":

    # folder_path = r"C:\Users\seba\Desktop\Proyectos\Repo_Leo\procesamiento_archivos\Data\Inputs\Inputs-2\Clientes_faltantes_lectura\D02-DSL FINISHING\12- December"
    # year = 2024
    # month = 12
    # total_files = 100 
    # processed_files = ""

    # df = process_weekly_files(folder_path, year, month, total_files, processed_files)
    # print(df.head())
    path_excel = r"C:\Users\seba\Desktop\Proyectos\Repo_Leo\procesamiento_archivos\services\output\2024_Noviembre_dataPDFs.xlsx"
    update_master_with_comparisons(path_excel)