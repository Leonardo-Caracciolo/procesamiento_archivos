
import os
import pytesseract
import fitz  # PyMuPDF para trabajar con PDFs
from PIL import Image, ImageEnhance, ImageFilter
import pandas as pd
import numpy as np
# import services.func_extrac_data as look_data  # Asegúrate de que esta ruta sea correcta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sys

from PyQt5.QtCore import QObject, pyqtSignal, QThread


def validation_data(path_file):
    """
    Indicamos resaltando con un color que archivos no fueron encontrados en sus carpetas
    """
    # Seleccionar archivo Excel
    excel_path = path_file

    # Abrir el archivo Excel con openpyxl
    try:
        wb = load_workbook(excel_path)
        ws = wb['Datos combinados']

        # Definir el color de fondo (amarillo)
        color_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        color_fill_format_data = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Insertar fórmulas en las columnas H y I
        for row in range(2, ws.max_row + 1):  # Asume que la fila 1 es el encabezado
            if (ws[f"D{row}"].value == "No se pudo obtener debido al formato del archivo" or
                ws[f"E{row}"].value == "No se pudo obtener debido al formato del archivo" or
                ws[f"F{row}"].value == "No se pudo obtener debido al formato del archivo" or
                ws[f"G{row}"].value == "No se pudo obtener debido al formato del archivo"):
                    for col in range(1, 12): # Recorrer todas las celdas en la fila
                        cell = ws.cell(row=row, column=col)
                        cell.fill = color_fill_format_data
            elif ws[f"K{row}"].value == "Archivo no encontrado":
                for col in range(1, 12): # Recorrer todas las celdas en la fila
                    cell = ws.cell(row=row, column=col)
                    cell.fill = color_fill

        ws_resumen = wb['Resumen']
        
        # Recorrer las filas del DataFrame
        for row in range(2, ws_resumen.max_row + 1):  # Asume que la fila 1 es el encabezado
            # Verificar las celdas en las columnas C, D, F y G
            if(ws_resumen[f"C{row}"].value == "No se pudo obtener debido al formato del archivo" or
                ws_resumen[f"D{row}"].value == "No se pudo obtener debido al formato del archivo" or
                ws_resumen[f"F{row}"].value == "No se pudo obtener debido al formato del archivo" or
                ws_resumen[f"G{row}"].value == "No se pudo obtener debido al formato del archivo"):
                # Si las celdas no contienen esos textos, agregar las fórmulas
                for col in range(1, 9): # Recorrer todas las celdas en la fila
                    cell = ws_resumen.cell(row=row, column=col)
                    cell.fill = color_fill_format_data
            
            elif (ws_resumen[f"C{row}"].value == "Archivo no encontrado" or
                ws_resumen[f"D{row}"].value == "Archivo no encontrado" or
                ws_resumen[f"F{row}"].value == "Archivo no encontrado" or
                ws_resumen[f"G{row}"].value == "Archivo no encontrado"):
                # Si las celdas no contienen esos textos, agregar las fórmulas
                for col in range(1, 9): # Recorrer todas las celdas en la fila
                    cell = ws_resumen.cell(row=row, column=col)
                    cell.fill = color_fill


        # Guardar cambios
        wb.save(excel_path)
        print("Excel modificado con exito.")
    except Exception as e:
        print("Error al identificar errores de OCR o archivos faltantes")

if __name__ == '__main__':
    ruta_excel = r"C:\Users\seba\Desktop\Proyectos\Repo_Leo\MAL-datos_combinados.xlsx"
    validation_data(ruta_excel)

# import os

# folder_path = r"C:\Users\seba\Downloads\InputsTesteo-20241213T020658Z-001\InputsTesteo\Carpeta falla\E15 - EL MICHOACANO PRODUCE INC\2024\12- December"
# year = '2024'  # Reemplaza con el año adecuado

# for file_name in os.listdir(folder_path):
#     file_path = os.path.join(folder_path, file_name)

#     # Verificar si el archivo existe y es un archivo
#     if not os.path.isfile(file_path):
#         continue

#     # Verificar si los primeros 8 dígitos del nombre del archivo son numéricos
#     if not file_name[:8].isdigit():
#         continue

#     # Verificar la longitud del nombre del archivo sin la extensión
#     nombre_sin_extension = os.path.splitext(file_name)[0]
#     if len(nombre_sin_extension) > 12:
#         continue

#     # Verificar los sufijos del nombre del archivo
#     if not any(file_name.endswith(suffix + ".pdf") for suffix in ["EDD", "941", f"{year}"]):
#         continue
    
#     if any(file_name.endswith(suffix + ".pdf") for suffix in ["EDD", "941"]):
#         caracter_nueve = file_name[8]
#         if caracter_nueve != ' ':
#             continue
#     elif any(file_name.endswith(suffix + ".pdf") for suffix in [f"{year}"]):
#         caracter_nueve = file_name[8]
#         if caracter_nueve != '.':
#             continue

#     # Aquí puedes continuar con tu lógica para procesar el archivo
#     print(f"Procesando archivo: {file_name}")
